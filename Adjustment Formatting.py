import win32com.client as win32
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def choose_file():
    """
    Opens a file dialog for the user to select a file.

    Returns:
        str: The full path to the selected file.
    """
    root = Tk()
    root.withdraw()
    file_path = askopenfilename(
        filetypes=[("Excel Files", "*.xlsx;*.xlsm")],
        title="Select an Excel file"
    )
    if not file_path:
        raise FileNotFoundError("No file was selected!")
    return file_path


def process_excel(file_path):
    """
    Processes the Excel file to sort and highlight data based on given conditions.
    """
    # Check if the file is a valid Excel file
    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError("Invalid file format. Please select a .xlsx or .xlsm file.")

    try:
        # Load workbook first to preserve formatting
        wb = load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        raise ValueError(
            f"Failed to open Excel file. Ensure it is not corrupted or open in another program. Error: {e}")

    # Load data using pandas without modifying column formatting
    df = pd.read_excel(file_path, engine='openpyxl')  # Explicitly using openpyxl to avoid format issues

    # Sort by "Code (Sell Branch)" (A-Z) and "ABS Cost" (high to low)
    df = df.sort_values(by=["Code (Sell Branch)", "ABS Cost"], ascending=[True, False])

    # Clear existing data in the sheet (except headers)
    ws.delete_rows(2, ws.max_row)

    # Write sorted data back to Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply short date format to the first column without modifying values
    date_style = NamedStyle(name="short_date")
    date_style.number_format = "MM/DD/YYYY"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "MM/DD/YYYY"

    # Apply comma 1,000 separator with decimal places to columns K, L, and N
    comma_style = NamedStyle(name="comma_sep")
    comma_style.number_format = "#,##0.00"
    for col in [11, 12, 14]:  # Columns K, L, and N are 11th, 12th, and 14th in 1-based index
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = "#,##0.00"

    # Define color fills
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Get column indices for relevant fields
    adj_type_col = df.columns.get_loc("Code (Adjustment Type)") + 1
    abs_cost_col = df.columns.get_loc("ABS Cost") + 1

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        adj_type = row[adj_type_col - 1].value
        abs_cost = row[abs_cost_col - 1].value

        if isinstance(abs_cost, (int, float)):
            if adj_type in ["B", "BP", "CC", "CL", "CYC", "DFC", "DIT", "FAB", "LIT", "SNP", "VC",
                            "WO"] and abs_cost > 300:
                row[abs_cost_col - 1].fill = green_fill
            elif adj_type == "DC":
                if abs_cost > 1000:
                    row[abs_cost_col - 1].fill = red_fill
                elif 300 < abs_cost <= 1000:
                    row[abs_cost_col - 1].fill = green_fill
            elif adj_type in ["HJ", "JWW", "MTD", "NA", "PO", "PPV", "TFT", "US", "PI"]:
                row[abs_cost_col - 1].fill = green_fill
            else:
                row[abs_cost_col - 1].fill = yellow_fill

    wb.save(file_path)
    print(f"Processing complete. File saved: {file_path}")
    return file_path


if __name__ == "__main__":
    try:
        file_path = choose_file()
        processed_file = process_excel(file_path)
        print(f"File saved: {processed_file}")
    except Exception as e:
        print(f"An error occurred: {e}")
